import csv
import requests
import numpy as np
from io import StringIO, BytesIO
from datetime import date, timedelta, datetime
from openpyxl import load_workbook
from global_vars import CBI_RATE_URL, NASDAQ_FIXED_DURATION_YIELD_URL, REIBOR_SERIES

class RiskFreeRate(float):
    def __new__(cls, start_date: date, end_date: date):
        rate = cls.get_annual_risk_free_interest_rate(start_date, end_date)
        return super().__new__(cls, rate)

    @staticmethod
    def latest_cbi_rate(series_id: int, as_of_date: date) -> float:
        response = requests.get(CBI_RATE_URL, params={
            "DagsFra": (as_of_date - timedelta(days=31)).isoformat(),
            "DagsTil": as_of_date.isoformat(),
            "TimeSeriesID": series_id,
            "Type": "csv",
        }, timeout=20)
        response.raise_for_status()
        rates = [(datetime.strptime(row[6], "%m/%d/%Y %I:%M:%S %p").date(), float(row[7]) / 100)
                 for row in csv.reader(StringIO(response.text), delimiter=";") if len(row) > 7]
        if not rates:
            raise ValueError(f"No CBI rate available on or before {as_of_date}")
        return max(rates, key=lambda item: item[0])[1]

    @classmethod
    def get_annual_risk_free_interest_rate(cls, start_date: date, end_date: date) -> float:
        tau = (end_date - start_date).days / 365
        if not 0 < tau <= 10:
            raise ValueError("Time between start_date and end_date must be between 0 and 10 years.")

        if tau <= 0.5:
            maturity, series_id = min(REIBOR_SERIES, key=lambda item: abs(item[0] - tau))
            rate = cls.latest_cbi_rate(series_id, start_date)
            return float(np.log1p(rate * maturity) / maturity)

        try:
            flv_rates = [cls.latest_cbi_rate(series_id, start_date) for series_id in (30110, 30111, 30112)]
        except ValueError:
            response = requests.get(NASDAQ_FIXED_DURATION_YIELD_URL, timeout=20)
            response.raise_for_status()
            rows = load_workbook(BytesIO(response.content), read_only=True, data_only=True)["Sheet1"].iter_rows(min_row=2, values_only=True)
            latest_row = max(
                (row for row in rows if row[1] and row[1].date() <= start_date),
                key=lambda row: row[1],
                default=None,
            )
            if latest_row is None:
                raise ValueError(f"No Nasdaq yield available on or before {start_date}")

            maturity_rates = [(maturity, rate) for maturity, rate in ((1, latest_row[4]), (5, latest_row[6]), (10, latest_row[3])) if rate is not None]
            if tau > maturity_rates[-1][0]:
                raise ValueError(f"No Nasdaq yield available for {tau:.2f} years on {start_date}")

            maturities = [maturity for maturity, _ in maturity_rates]
            rates = [np.log1p(rate) for _, rate in maturity_rates]
            if tau < 1:
                short_rate = cls.latest_cbi_rate(16, start_date)
                maturities.insert(0, 0.5)
                rates.insert(0, np.log1p(short_rate * 0.5) / 0.5)
            return float(np.interp(tau, maturities, rates))

        short_rate = cls.latest_cbi_rate(16, start_date)
        maturities = np.array([0.5, 3, 5, 10])
        rates = np.array([
            np.log1p(short_rate * 0.5) / 0.5,
            *np.log1p(flv_rates),
        ])
        log_discounts = -maturities * rates
        return float(-np.interp(tau, maturities, log_discounts) / tau)