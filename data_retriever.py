import csv
import requests
import numpy as np
from io import StringIO
from datetime import date, timedelta, datetime
import yfinance as yf

from black_scholes import call_option_price, put_option_price
from global_vars import ICELANDIC_MARKET_CLOSED_DATES, CBI_RATE_URL, NASDAQ_FIXED_DURATION_YIELD_URL, REIBOR_SERIES

#! mögulega færa adjust date föllinn eh annað og gera ráð fyrir sanatized dateum
#! change this to a class maybe

def get_stock_info(stock_name: str, start_date: date, end_date: date, price_column: str = "Close"):
    stock_ticker = yf.Ticker(stock_name)
    hist = stock_ticker.history(period="max", auto_adjust=False)

    if start_date < (first_date := hist.index.min().date()):
        start_date = first_date

    if end_date > (last_date := hist.index.max().date()):
        end_date = last_date

    hist = hist[(hist.index.date >= start_date) & (hist.index.date <= end_date)]
    return hist[[price_column, "Volume"]].rename_axis("Date").reset_index()

def get_current_stock_price(stock_name: str, current_date: date, price_column: str = "Close"):
    current_date = adjust_date(current_date)
    end_date = current_date + timedelta(days=1)
    current_stock_info = get_stock_info(stock_name=stock_name, start_date=current_date, end_date=end_date, price_column=price_column)
    return current_stock_info.loc[current_stock_info["Date"].dt.date == current_date, price_column].iloc[0]

    
def get_volatility(stock_name: str, lookback_days: int = 60, current_date: date = date.today(), price_column: str = "Close") -> float:
    lookback_days = max(30, lookback_days)
    stock_ticker = yf.Ticker(stock_name)
    hist = stock_ticker.history(period="max", auto_adjust=True)
    start_calc_vol_date = adjust_date(current_date - timedelta(days=lookback_days))
    past_stock_prices = hist.loc[(hist.index.date >= start_calc_vol_date) & (hist.index.date <= current_date), price_column].to_numpy()
    """
    start_to_end_delta = end_date - start_date
    lookback_days = max(30, start_to_end_delta.days)
    start_calc_vol_date = current_date - timedelta(days=lookback_days) 
    past_stock_info = get_stock_info(stock_name=stock_name, start_date=start_calc_vol_date, end_date=current_date, price_column=price_column)
    """
    log_returns = np.log(past_stock_prices[1:] / past_stock_prices[:-1])
    std_of_log_returns = np.std(log_returns, ddof=1)
    trading_periods = np.sqrt(252) # 252 is the number of trading days in a year
    return std_of_log_returns * trading_periods


def latest_cbi_rate(series_id: int, as_of_date: date) -> float:
    response = requests.get(
        CBI_RATE_URL,
        params={
            "DagsFra": (as_of_date - timedelta(days=31)).isoformat(),
            "DagsTil": as_of_date.isoformat(),
            "TimeSeriesID": series_id,
            "Type": "csv",
        }, timeout=20)
    response.raise_for_status()
    rates = [
        (
            datetime.strptime(row[6], "%m/%d/%Y %I:%M:%S %p").date(),
            float(row[7]) / 100,
        )
        for row in csv.reader(StringIO(response.text), delimiter=";")
        if len(row) > 7
    ]
    if not rates:
        raise ValueError(f"No CBI rate available on or before {as_of_date}")
    return max(rates, key=lambda item: item[0])[1]


def get_annual_risk_free_interest_rate(T: float, t: float, as_of_date: date) -> float:
    """Return the continuously compounded annual ISK rate."""
    tau = T - t
    if not 0 < tau <= 10:
        raise ValueError("T - t must be between 0 and 10 years.")

    if tau <= 0.5:
        maturity, series_id = min(
            REIBOR_SERIES,
            key=lambda item: abs(item[0] - tau),
        )
        rate = latest_cbi_rate(series_id, as_of_date)
        return float(np.log1p(rate * maturity) / maturity)

    try:
        flv_rates = [
            latest_cbi_rate(series_id, as_of_date)
            for series_id in (30110, 30111, 30112)
        ]
    except ValueError:
        from io import BytesIO
        from openpyxl import load_workbook

        response = requests.get(
            NASDAQ_FIXED_DURATION_YIELD_URL,
            timeout=20,
        )
        response.raise_for_status()
        rows = load_workbook(
            BytesIO(response.content), read_only=True, data_only=True
        )["Sheet1"].iter_rows(min_row=2, values_only=True)
        latest_row = max(
            (row for row in rows if row[1] and row[1].date() <= as_of_date),
            key=lambda row: row[1],
            default=None,
        )
        if latest_row is None:
            raise ValueError(f"No Nasdaq yield available on or before {as_of_date}")

        maturity_rates = [(maturity, rate) for maturity, rate in ((1, latest_row[4]), (5, latest_row[6]), (10, latest_row[3])) if rate is not None]
        if tau > maturity_rates[-1][0]:
            raise ValueError(f"No Nasdaq yield available for {tau:.2f} years on {as_of_date}")

        maturities = [maturity for maturity, _ in maturity_rates]
        rates = [np.log1p(rate) for _, rate in maturity_rates]
        if tau < 1:
            short_rate = latest_cbi_rate(16, as_of_date)
            maturities.insert(0, 0.5)
            rates.insert(0, np.log1p(short_rate * 0.5) / 0.5)
        return float(np.interp(tau, maturities, rates))

    short_rate = latest_cbi_rate(16, as_of_date)
    maturities = np.array([0.5, 3, 5, 10])
    rates = np.array([
        np.log1p(short_rate * 0.5) / 0.5,
        *np.log1p(flv_rates),
    ])
    log_discounts = -maturities * rates
    return float(-np.interp(tau, maturities, log_discounts) / tau)

def adjust_date(date: date):
    """Adjust weekend dates to the last trading day"""
    if date in ICELANDIC_MARKET_CLOSED_DATES or date.isoweekday() == 6: # 6 is a saturday
        return adjust_date(date - timedelta(days=1))
    if date.isoweekday() == 7: # 7 is sunday
        return adjust_date(date - timedelta(days=2))
    return date

def get_T_and_t(start_date: date, end_date: date, current_date: date = date.today()):
    start_to_end_delta = end_date - start_date
    start_to_current_delta = current_date - start_date 
    return start_to_end_delta.days / 365, start_to_current_delta.days / 365

def get_info_for_option_price(stock_name: str, start_date: date, end_date: date, current_date: date = date.today()):
    start_date, end_date, current_date = adjust_date(start_date), adjust_date(end_date), adjust_date(current_date)
    T, t = get_T_and_t(start_date=start_date, end_date=end_date, current_date=current_date)
    stock_price = get_current_stock_price(stock_name=stock_name, current_date=current_date)
    volatility = get_volatility(stock_name=stock_name, current_date=current_date)
    interest_rate = get_annual_risk_free_interest_rate(T=T, t=t, as_of_date=current_date)
    if volatility <= 0:
        print(f"{stock_name} - Zero or lower volatility: {volatility}, T: {T}, t: {t}, interest_rate: {interest_rate}")
    return stock_price, T, t, volatility, interest_rate


def call_option_stock_price(stock_name: str, strike_price: float, start_date: date, end_date: date, current_date: date = date.today()):
    stock_price, T, t, volatility, interest_rate = get_info_for_option_price(stock_name=stock_name, start_date=start_date, end_date=end_date, current_date=current_date)
    return call_option_price(S=stock_price, K=strike_price, T=T, t=t, volatility=volatility, r=interest_rate)


def put_option_stock_price(stock_name: str, strike_price: float, start_date: date, end_date: date, current_date: date = date.today()):
    stock_price, T, t, volatility, interest_rate = get_info_for_option_price(stock_name=stock_name, start_date=start_date, end_date=end_date, current_date=current_date)
    return put_option_price(S=stock_price, K=strike_price, T=T, t=t, volatility=volatility, r=interest_rate)
